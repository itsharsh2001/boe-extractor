import pandas as pd
import numpy as np
import aspose.pdf as ap
import PyPDF2
from openpyxl import load_workbook
import os
from datetime import datetime
import tempfile


def Extract(input_file: str) -> pd.DataFrame | None:
    if not os.path.exists(input_file):
        return None

    tempdir = tempfile.TemporaryDirectory(
        ignore_cleanup_errors=True, prefix="Grant_Thornton"
    )
    output_dir = tempdir.name

    file_name = os.path.basename(input_file)

    try:
        temp_excel_file = convert_pdf_to_excel(
            input_file=input_file, output_dir=output_dir
        )
    except Exception as e:
        print(f"Error converting pdf to excel for file {file_name}. Full Error:{e}")
        tempdir.cleanup()
        return None

    try:
        newdf = extract_data_from_excel(input_file=temp_excel_file)
        if len(newdf) == 0:
            print(f"Unable to extract data from file {file_name}.")
            tempdir.cleanup()
            return None
    except Exception as e:
        print(f"Error extracting data for file {file_name}. Full Error:{e}")
        tempdir.cleanup()
        return None

    tempdir.cleanup()
    return newdf


def Extract_all(input_directory: str) -> tuple[str, list[str], list[str]] | None:
    if not os.path.isdir(input_directory):
        raise OSError(f"Directory does not exist: {input_directory}")

    file_list = [
        os.path.join(input_directory, file)
        for file in os.listdir(input_directory)
        if os.path.splitext(file)[1] == ".pdf"
    ]

    if len(file_list) == 0:
        print("No PDF Files found in the folder")
        return None

    finaldf = pd.DataFrame()

    correct_files = []
    incorrect_files = []

    for filename in file_list:
        newdf = Extract(input_file=filename)

        if newdf is None:
            print(f"Unable to extract for {filename}")
            incorrect_files.append(filename)
            continue

        correct_files.append(filename)

        finaldf = pd.concat([finaldf, newdf])

    if len(finaldf) == 0:
        print("Unable to extract any data")
        return None

    try:
        output_file = format_and_save_df(df=finaldf, input_dir=input_directory)
        print(output_file)
    except Exception as e:
        print(e)
        print("Unable to save extracted any data")
        return None

    return (output_file, correct_files, incorrect_files)


def copy_sheet(source_sheet, target_wb, sheet_title):
    target_sheet = target_wb.create_sheet(sheet_title)

    # Copy the data from source to target sheet
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(row)

    # Copy images, if any
    # for image in source_sheet._images:
    #    target_sheet.add_image(image)

    # Copy charts, if any
    # for chart in source_sheet._charts:
    #    target_sheet.add_chart(chart)


def format_and_save_df(df: pd.DataFrame, input_dir: str) -> str:
    filename = f"output {str(datetime.now())[:19]}.xlsx"
    filename = os.path.join(input_dir, filename.replace(":", ";"))
    while os.path.exists(filename):
        filename = f"output {str(datetime.now())[:19]}.xlsx"
        filename = os.path.join(input_dir, filename.replace(":", ";"))

    writer = pd.ExcelWriter(filename, engine="xlsxwriter")
    df.to_excel(writer, index=False, startrow=2, header=False)
    workbook = writer.book
    worksheet = writer.sheets["Sheet1"]

    header_format = workbook.add_format(  # type: ignore
        {
            "bold": True,
            "text_wrap": True,
            "valign": "centre",
            "fg_color": "#BFBFBF",
            "border": 1,
        }
    )

    for col_num, value in enumerate(df.columns.values):
        worksheet.write(1, col_num, value, header_format)

    writer.close()

    return filename


def convert_pdf_to_excel(input_file: str, output_dir: str) -> str:
    output_file = os.path.join(
        output_dir, f"{os.path.splitext(os.path.basename(input_file))[0]}.xlsx"
    )
    print(output_file,'output file is here')
    temp_df = pd.DataFrame()
    temp_df.to_excel(output_file)
    pdf_reader = PyPDF2.PdfReader(input_file)
    num_pages = (
        len(pdf_reader.pages) - 3
    )  # because last 3 pages does not contain any useful information
    temppdf_filepath = os.path.join(
        output_dir, f"{os.path.splitext(os.path.basename(input_file))[0]}_temp_file.pdf"
    )
    tempexcel_filepath = os.path.join(
        output_dir,
        f"{os.path.splitext(os.path.basename(input_file))[0]}_temp_file.xlsx",
    )
    merged_wb = load_workbook(filename=output_file)

    for i in range(0, num_pages, 4):
        pdf_writer = PyPDF2.PdfWriter()

        for j in range(i, i + 4):
            if j < num_pages:
                pdf_writer.add_page(pdf_reader.pages[j])
            else:
                break

        with open(temppdf_filepath, "wb") as f:
            pdf_writer.write(f)

        document = ap.Document(temppdf_filepath)  # type: ignore
        document.save(tempexcel_filepath, ap.ExcelSaveOptions())

        temp_wb = load_workbook(filename=tempexcel_filepath)
        count = 0
        if i == 0:
            merged_wb.remove(merged_wb["Sheet1"])
        for sheets in temp_wb.sheetnames:
            count += 1
            copy_sheet(temp_wb[sheets], merged_wb, f"Sheet{i+count}")

    merged_wb.save(output_file)

    try:
        os.remove(temppdf_filepath)
    except:
        pass
    try:
        os.remove(tempexcel_filepath)
    except:
        pass

    return output_file


def extract_data_from_excel(input_file: str) -> pd.DataFrame:
    itemdf = pd.DataFrame()

    sheet_names = pd.ExcelFile(input_file).sheet_names
    sheetdf = pd.DataFrame(columns=["Part", "Sheet Name"])
    for index, i in enumerate(sheet_names):
        new_row = []
        temp_df = pd.read_excel(input_file, sheet_name=i)
        ls = list(temp_df.iloc[6, :])
        while np.nan in ls:
            ls.remove(np.nan)
        ls = list(map(str, ls))
        string = "".join(ls)
        new_row.append(string.split("-")[1].lstrip().rstrip())
        new_row.append(i)
        sheetdf.loc[index, :] = new_row

    df = pd.read_excel(
        input_file, sheet_name=sheetdf[sheetdf["Part"] == "I"]["Sheet Name"][0]
    )
    i = 0
    while df.iloc[0, i] in [np.nan, "nan", float(np.nan)]:
        i += 1
    port_code = str(df.iloc[0, i]).split(sep="\n")[-1]
    BOE_No = str(df.iloc[0, i + 1]).split(sep="\n")[-1]
    BOE_Date = str(df.iloc[0, i + 2]).split(sep="\n")[-1]
    GSTIN = str(df.iloc[2, i + 1])[:-2]
    number_of_invoices = int(df.iloc[4, i + 1])
    a = list(map(str, list(df.iloc[9, :])))
    while "nan" in a:
        a.remove("nan")
    country_of_origin = a[1]
    invoicecolumns = ["INVOICESN", "Invoice No.", "Invoice Value", "Invoice Currency"]

    for i in df.columns:
        a = list(df.loc[:, i])
        if "2.INVOICE NO" in a:
            invoice_start = a.index("2.INVOICE NO")
            invoicedf = pd.read_excel(
                input_file,
                sheet_name=sheetdf[sheetdf["Part"] == "I"]["Sheet Name"][0],
                skiprows=int(invoice_start) + 1,
                nrows=number_of_invoices,
            )
            invoicedf = invoicedf.iloc[:, list(invoicedf.columns).index("1.S.NO") :]
            invoicedf = invoicedf.dropna(axis=1, how="all")
            break
    else:
        invoicedf = pd.DataFrame()

    if invoicedf.shape[1] == 4:
        invoicedf.columns = invoicecolumns
    else:
        newdf = pd.DataFrame(columns=invoicecolumns)
        for i in range(len(invoicedf)):
            new_row = list(invoicedf.iloc[i, :])
            while np.nan in new_row:
                new_row.remove(np.nan)
            if len(new_row) >= 4:
                new_row = new_row[:4]
            else:
                while len(new_row) != 4:
                    new_row.append("")
            newdf.iloc[i, :] = new_row  # type: ignore
        invoicedf = newdf.copy()

    df = pd.read_excel(input_file, sheet_name="Sheet1")

    # Exchange Rate
    invoicedf["Exchange Rate"] = float(np.nan)
    for col in df.columns:
        column = list(df[col])
        if "EXCHANGE RATE" in column:
            index = column.index("EXCHANGE RATE") + 1
            end = column.index("5.CONTAINER NUMBER")
            column = column[index:end]
            column = list(map(str, column))
            for i in invoicedf.index:
                currency = invoicedf.loc[i, "Invoice Currency"]
                for val in column:
                    if currency in val:
                        exchange_rate = val.split("=")[-1].split("INR")[0]
                        try:
                            exchange_rate = float(exchange_rate)
                        except:
                            pass
                        invoicedf.loc[i, "Exchange Rate"] = exchange_rate
                        break
                else:
                    column = list(df.iloc[:, list(df.columns).index(col) + 1])
                    column = column[index:end]
                    column = list(map(str, column))
                    for val in column:
                        if currency in val:
                            exchange_rate = val.split("=")[-1].split("INR")[0]
                            try:
                                exchange_rate = float(exchange_rate)
                            except:
                                pass
                            invoicedf.loc[i, "Exchange Rate"] = exchange_rate
                            break
            break

    # HAWB No. & Date
    row = list(df.iloc[:, 0]).index("1.IGM NO") + 1
    hawb_mawb_df = pd.read_excel(input_file, sheet_name="Sheet1", skiprows=row, nrows=1)
    hawb_mawb_df = hawb_mawb_df.dropna(how="all", axis=1)
    a = list(hawb_mawb_df.iloc[0, list(hawb_mawb_df.columns).index("8.HAWB NO") :])
    invoicedf["HAWB No."] = a[0]
    invoicedf["HAWB Date"] = a[1]

    # MAWB No. & Date
    b = list(hawb_mawb_df.iloc[0, list(hawb_mawb_df.columns).index("6.MAWB NO") :])
    invoicedf["MAWB No."] = b[0]
    invoicedf["MAWB Date"] = b[1]

    "1.IMPORTER Details"
    new_row = list(df.iloc[:, 0])
    for index, value in enumerate(new_row):
        if "1.IMPORTER" in str(value):
            new_row = index + 1
            importer_df = pd.read_excel(
                input_file, sheet_name="Sheet1", skiprows=new_row
            )
            col = list(importer_df.iloc[:, 0])
            while np.nan in col:
                col.remove(np.nan)
            address = []
            for index, value in enumerate(col):
                value = str(value)
                if index == 0:
                    invoicedf["Importer Name"] = value

                if "AD CODE" in value:
                    break

                address.append(value)

            invoicedf["Importer No"] = ""
            invoicedf["Importer Address"] = ", ".join(address)
            break
    else:
        invoicedf["Importer No"] = ""
        invoicedf["Importer Name"] = ""
        invoicedf["Importer Address"] = ""

    # AD Code
    try:
        new_row = list(df.iloc[:, 0]).index("AD CODE") + 1
        adcode_df = pd.read_excel(input_file, sheet_name="Sheet1", skiprows=new_row)
        ad_code = str(list(adcode_df.columns)[1])
        if "Unnamed" in ad_code:
            invoicedf["AD Code"] = ""
        else:
            invoicedf["AD Code"] = str(list(adcode_df.columns)[1])
    except:
        invoicedf["AD Code"] = ""

    # Duties
    new_row = list(df.iloc[:, 0]).index("1.BCD") + 1
    duties_df = pd.read_excel(input_file, sheet_name="Sheet1", skiprows=new_row)

    ls1 = ["Total BCD", "Total SWS", "Total IGST"]
    ls2 = ["1.BCD", "3.SWS", "7.IGST"]

    for i, v in zip(ls1, ls2):
        try:
            invoicedf[i] = duties_df.loc[0, v]
        except:
            invoicedf[i] = ""

    duties_df.columns = duties_df.iloc[1, :]
    duties_df = duties_df.iloc[2:, :]
    duties_df = duties_df.reset_index(drop=True)
    print(duties_df.columns)
    ls1 = ["Total Duty", "Interest", "Penalty", "Fine"]
    ls2 = ["14.TOTAL DUTY", "15.INT", "16.PNLTY", "17.FINE"]

    try:
        val = duties_df.loc[0, "14.TOTAL DUTY"]
        invoicedf["Total Duty"] = val
        row = list(duties_df.iloc[0, :])
        print(val)
        print(row)
        while np.nan in row:
            row.remove(np.nan)
        index = row.index(val)
        if index + 4 != len(row):
            row = row[index + 1 :]
            invoicedf["Interest"] = row[0]
            invoicedf["Penalty"] = row[1]
            invoicedf["Fine"] = row[2]

    except:
        for i, v in zip(ls1, ls2):
            try:
                invoicedf[i] = duties_df.loc[0, v]
            except:
                invoicedf[i] = ""

    row = list(duties_df.iloc[0, :])
    while np.nan in row:
        row.remove(np.nan)

    # OOC
    try:
        row = list(df.iloc[:, 0]).index("1.EVENT") + 1
        df = pd.read_excel(input_file, sheet_name="Sheet1", skiprows=row)
        invoicedf["OOC"] = df.iloc[
            list(df.iloc[:, 0]).index("OOC"), list(df.columns).index("2.DATE")
        ]
    except:
        invoicedf["OOC"] = np.nan

    invoicedf["Port Code"] = port_code
    invoicedf["BOE No."] = BOE_No
    invoicedf["BOE Date"] = BOE_Date
    invoicedf["GSTIN Details"] = GSTIN
    invoicedf["Country of Origin"] = country_of_origin
    invoicedf["Invoice Date"] = np.nan
    invoicedf["Supplier Name"] = np.nan
    invoicedf["Third Party"] = np.nan
    invoicedf["TOI"] = np.nan
    invoicedf["Total Freight"] = float(np.nan)
    invoicedf["Freight Currency"] = np.nan
    invoicedf["Total Insurance"] = float(np.nan)
    invoicedf["Insurance Currency"] = np.nan
    invoicedf["Total Misc Charges"] = float(np.nan)
    invoicedf["Port Code & BOE No."] = (
        invoicedf["Port Code"] + " & " + invoicedf["BOE No."].astype(str)
    )
    invoicedf = invoicedf.set_index("INVOICESN", drop=True)

    count = 0
    for sheet in sheetdf[sheetdf["Part"] == "II"]["Sheet Name"]:
        df = pd.read_excel(input_file, sheet_name=sheet, skiprows=8)

        first_row = list(df.loc[0, :])
        while np.nan in first_row:
            first_row.remove(np.nan)
        invoice_serial_number = int(first_row[0])

        if count != invoice_serial_number:
            count += 1
            second_row = list(df.loc[1, :])
            while np.nan in second_row:
                second_row.remove(np.nan)
            invoicedf.loc[count, "Invoice Date"] = second_row[0]

            for i, v in enumerate(df.columns):
                if "15.Term" in list(df.loc[:, v]):
                    invoicedf.loc[count, "TOI"] = df.iloc[
                        list(df.loc[:, v]).index("15.Term"), i + 1
                    ]

            for i, v in enumerate(df.columns):
                if "1.INV VALUE" in list(df.loc[:, v]):
                    newdf = df.iloc[list(df.loc[:, v]).index("1.INV VALUE") :, :].copy()
                    newdf.columns = newdf.iloc[0, :]
                    newdf = newdf.iloc[1:, :]
                    newdf = newdf.reset_index(drop=True)
                    invoicedf.loc[count, "Total Freight"] = float(newdf["2.FREIGHT"][0])
                    invoicedf.loc[count, "Freight Currency"] = newdf["2.FREIGHT"][1]
                    invoicedf.loc[count, "Total Insurance"] = float(
                        newdf["3.INSURANCE"][0]
                    )
                    invoicedf.loc[count, "Insurance Currency"] = newdf["3.INSURANCE"][1]
                    break

            for i, v in enumerate(df.columns):
                if "13.MISC CHARGE" in list(df.loc[:, v]):
                    newdf = df.iloc[
                        list(df.loc[:, v]).index("13.MISC CHARGE") :, :
                    ].copy()
                    newdf.columns = newdf.iloc[0, :]
                    newdf = newdf.iloc[1:, :]
                    newdf = newdf.reset_index(drop=True)
                    invoicedf.loc[count, "Total Misc Charges"] = float(
                        newdf["13.MISC CHARGE"][0]
                    )

            for i, v in enumerate(df.index):
                rows = df.loc[v, :]
                for j in rows:
                    if "3.SUPPLIER" in str(j):
                        df.columns = df.iloc[i]
                        df = df.iloc[i + 1 :, :]
                        df.reset_index(drop=True, inplace=True)
                        break
                else:
                    continue
                break

            for i, v in enumerate(df.columns):
                if "3.SUPPLIER" in str(v):
                    for j in range(-2, 3):
                        if i + j < 0 or i + j >= len(df.columns):
                            continue
                        if str(df.iloc[0, i + j]) != str(np.nan):
                            if "\n" in df.iloc[0, i + j]:
                                invoicedf.loc[count, "Supplier Name"] = df.iloc[
                                    0, i + j
                                ].split("\n")[0]
                            else:
                                invoicedf.loc[count, "Supplier Name"] = df.iloc[
                                    0, i + j
                                ]
                            break

                    break

            for i, v in enumerate(df.columns):
                if "4.THIRD" in str(v):
                    for j in range(-2, 3):
                        if i + j < 0 or i + j >= len(df.columns):
                            continue
                        if str(df.iloc[0, i + j]) != str(np.nan):
                            if "\n" in df.iloc[0, i + j]:
                                invoicedf.loc[count, "Third Party"] = df.iloc[
                                    0, i + j
                                ].split("\n")[0]
                            else:
                                invoicedf.loc[count, "Third Party"] = df.iloc[0, i + j]
                            break
                    break
            pass

        df = pd.read_excel(input_file, sheet_name=sheet)
        start = len(df) - list(df.iloc[:, 0])[::-1].index("1.S NO.")
        df = pd.read_excel(input_file, sheet_name=sheet, skiprows=start)
        df = df.iloc[: list(df.iloc[:, 0]).index("GLOSSARY"), :]
        df = df.dropna(axis=1, how="all")
        df = df.fillna(value="")
        df = df.astype(str)

        while "" in list(df.iloc[:, 0]):
            a = list(df.iloc[:, 0])
            start = a.index("")
            for i in range(start + 1, len(a)):
                if a[i] == "":
                    continue
                else:
                    end = i
                    break
            else:
                end = len(a)
            for i in range(len(df.columns)):
                df.iloc[start - 1, i] = "".join(df.iloc[start - 1 : end, i])
            df = df.drop(list(range(start, end)))
            df = df.reset_index(drop=True)

        itemcolumn = [
            "ITEMSN",
            "HSN Code",
            "DESCRIPTION",
            "Unit Price",
            "Quantity",
            "UQC",
            "Amount",
        ]

        if df.shape[1] == 7:
            df.columns = itemcolumn
        else:
            newdf = pd.DataFrame(columns=itemcolumn)
            for i in range(len(df)):
                new_row = list(df.iloc[i, :])
                while "" in new_row:
                    new_row.remove("")
                if len(new_row) >= 7:
                    new_row = new_row[:7]
                else:
                    while len(new_row) != 7:
                        new_row.append("")
                newdf = pd.concat(
                    [newdf, pd.DataFrame([new_row], columns=itemcolumn)],
                    ignore_index=True,
                )
            df = newdf.copy()

        pattern = r"(\d{4}[A-Z]\d{3})"

        df["Style"] = df["DESCRIPTION"].str.extract(pattern)

        # for i in ["STYLE NO.", "STYLE:", "S/NO:"]:
        #     if i in df["DESCRIPTION"][0]:
        #         df["Style"] = df["DESCRIPTION"].apply(
        #             lambda x: x.split(i)[1][:8]
        #         )
        #         break
        # else:
        #     df["Style"] = np.nan

        if "CLR:" in df["DESCRIPTION"][0]:
            df["Color"] = df["DESCRIPTION"].apply(lambda x: x.split("CLR:")[1][:3])
        else:
            df["Color"] = np.nan

        df["INVOICESN"] = count
        df["Quantity"] = df["Quantity"].astype("float")

        # count += 1

        itemdf = pd.concat([itemdf, df])

    itemdf["Total Quantity"] = itemdf.groupby("INVOICESN")["Quantity"].transform("sum")

    # itemdf["Total Quantity"] = df["Quantity"].sum()
    # print(f"itemdf: {itemdf}")
    invoicedf = invoicedf.reset_index(names="INVOICESN")
    invoicedf["INVOICESN"] = invoicedf["INVOICESN"].astype(int)

    dutycolumns = [
        "INVOICESN",
        "ITEMSN",
        "Assessed Value",
        "BCD Rate",
        "SWS Rate",
        "IGST Rate",
        "BCD",
        "SWS",
        "IGST",
    ]
    dutiesdf = pd.DataFrame(columns=dutycolumns)

    for sheet in sheetdf[sheetdf["Part"] == "III"]["Sheet Name"]:
        newdf = pd.DataFrame(columns=dutycolumns)
        df = pd.read_excel(input_file, sheet_name=sheet, skiprows=7)
        for i, v in enumerate(df.columns):
            if "1.INVSNO" in list(df.loc[:, v]):
                break
        else:
            continue
        count = 0
        while "1.INVSNO" in list(df.iloc[:, i]):
            new_row = []
            row = list(df.iloc[:, i]).index("1.INVSNO")
            df.columns = df.iloc[row, :]
            df = df.iloc[row + 1 :, :]
            new_row.append(df["1.INVSNO"].iloc[0])
            new_row.append(df["2.ITEMSN"].iloc[0])
            a = list(
                map(
                    lambda x: x[:7] if len(x) > 7 else x,
                    df.iloc[:, i].astype(str),
                )
            )
            row = a.index("23.PROD")
            df.columns = df.iloc[row, :]
            df = df.iloc[row + 1 :, :]
            row = list(df.iloc[0, :])
            while np.nan in row:
                row.remove(np.nan)

            ls = []
            for j in row:
                try:
                    ls.append(float(j))
                except:
                    continue
            row = ls

            if len(row) == 2:
                new_row.append(row[0])
            else:
                a = list(
                    map(
                        lambda x: x[:9] if len(x) > 9 else x,
                        df.columns.astype(str),
                    )
                ).index("29.ASSESS")
                try:
                    new_row.append(df.columns[a].split("\n")[1])
                except:
                    if len(row) != 0:
                        new_row.append(row[0])
                    else:
                        new_row.append(np.nan)

            row = list(map(lambda x: x.split("\n"), df.iloc[:, i].astype(str)))
            for m, x in enumerate(row):
                if "DUTY" in x:
                    dutycolumn = "\n".join(x)
                    row = m
                    break

            df.columns = list(
                map(lambda x: x.split("\n")[0], df.iloc[row, :].astype(str))
            )
            df = df.iloc[row + 1 :, :]
            dutycolumn = "DUTY"

            a = df.iloc[:4, :].copy()
            unwanted_values = [
                np.nan,
                float(np.nan),
                0,
                0.0,
                "0",
                "0.0",
                "nan",
                float("nan"),
            ]
            # for x in a:
            #     a[x] = a[x].apply(lambda y: str(y).split("\n"))
            for x in range(a.shape[1]):
                for z in range(a.shape[0]):
                    a.iloc[z, x] = str(a.iloc[z, x]).split("\n")

            for x in a.index:
                if "Rate" in a.loc[x, dutycolumn]:
                    rateindex = int(a.loc[x, dutycolumn].index("Rate"))
                    rate_row = []
                    for j in a.loc[x, :]:
                        try:
                            if len(j) > 1:
                                val = j[rateindex]
                            else:
                                val = j[0]
                            if val not in unwanted_values:
                                rate_row.append(float(val))
                        except:
                            continue
                    # BCD
                    if len(rate_row) >= 1:
                        bcd = rate_row[0]
                    else:
                        bcd = 0
                    # SWS
                    if len(rate_row) < 3:
                        sws = 0
                    else:
                        sws = rate_row[1]
                    # IGST
                    if len(rate_row) >= 3:
                        igst = rate_row[2]
                    elif len(rate_row) == 2:
                        igst = rate_row[1]
                    else:
                        igst = 0
                    new_row.extend([bcd, sws, igst])
                    # new_row.append(a.loc[x, "1. BCD"][rateindex])
                    # new_row.append(a.loc[x, "3.SWS"][rateindex])
                    # new_row.append(a.loc[x, "5.IGST"][rateindex])
                    break
            for x in a.index:
                if "Amount" in a.loc[x, dutycolumn]:
                    amountindex = a.loc[x, dutycolumn].index("Amount")
                    amount_row = []
                    for j in a.loc[x, :]:
                        try:
                            if len(j) > 1:
                                val = j[amountindex]
                            else:
                                val = j[0]
                            if val not in unwanted_values:
                                amount_row.append(float(val))
                        except:
                            continue
                    # BCD
                    if len(amount_row) >= 1:
                        bcd = amount_row[0]
                    else:
                        bcd = 0
                    # SWS
                    if len(amount_row) < 3:
                        sws = 0
                    else:
                        sws = amount_row[1]
                    # IGST
                    if len(amount_row) >= 3:
                        igst = amount_row[2]
                    elif len(amount_row) == 2:
                        igst = amount_row[1]
                    else:
                        igst = 0
                    new_row.extend([bcd, sws, igst])

                    # new_row.append(a.loc[x, "1. BCD"][amountindex])
                    # new_row.append(a.loc[x, "3.SWS"][amountindex])
                    # new_row.append(a.loc[x, "5.IGST"][amountindex])
                    break
            newdf.loc[newdf.shape[0], :] = new_row

        dutiesdf = pd.concat([dutiesdf, newdf])

    # for i in ["ITEMSN", "INVOICESN"]:
    #     itemdf[i] = itemdf[i].astype(int)
    #     dutiesdf[i] = dutiesdf[i].astype(int)

    itemdf["ITEMSN"] = itemdf["ITEMSN"].astype(int)
    itemdf["INVOICESN"] = itemdf["INVOICESN"].astype(int)
    dutiesdf["ITEMSN"] = dutiesdf["ITEMSN"].astype(int)
    dutiesdf["INVOICESN"] = dutiesdf["INVOICESN"].astype(int)

    df = pd.merge(itemdf, dutiesdf, how="left", on=["INVOICESN", "ITEMSN"])

    newdf = pd.merge(
        invoicedf, df, how="left", left_on="INVOICESN", right_on="INVOICESN"
    )

    for i in [
        "Assessed Value",
        "BCD",
        "SWS",
        "IGST",
        "Quantity",
        "Total Freight",
        "Total Quantity",
        "Total Insurance",
        "Total Misc Charges",
        "Unit Price",
        "BCD Rate",
        "SWS Rate",
        "IGST Rate",
    ]:
        newdf[i] = newdf[i].astype(float)

    # if len(newdf["Invoice No."].unique()) != 1:
    #     for i in newdf.index:
    #         invoice_no = newdf.loc[i, "Invoice No."]
    #         total_quantity = newdf[newdf["Invoice No."] == invoice_no]["Quantity"].sum()
    #         quantity = newdf.loc[i, "Quantity"]
    #         for j in ["Freight", "Insurance", "Misc Charges"]:
    #             newdf.loc[i, j] = round(newdf.loc[i, f"Total {j}"] * quantity / total_quantity, 2)  # type: ignore
    # else:
    for i in ["Freight", "Insurance", "Misc Charges"]:
        newdf[i] = round(
            newdf[f"Total {i}"] * newdf["Quantity"] / newdf["Total Quantity"],
            2,
        )
        if len(newdf["Invoice No."].unique()) == 1:
            difference = newdf.loc[0, f"Total {i}"] - newdf[i].sum()
            newdf.loc[len(newdf) - 1, i] += difference

    newdf["HSN Code"] = newdf["HSN Code"].astype(int)
    newdf["IGST Assessed Value"] = newdf["Assessed Value"] + newdf["BCD"] + newdf["SWS"]

    column_order = [
        "Importer No",
        "Importer Name",
        "Importer Address",
        "AD Code",
        "Supplier Name",
        "Third Party",
        "Country of Origin",
        "HAWB No.",
        "HAWB Date",
        "MAWB No.",
        "MAWB Date",
        "Port Code",
        "BOE No.",
        "Port Code & BOE No.",
        "BOE Date",
        "Total BCD",
        "Total SWS",
        "Total IGST",
        "Total Duty",
        "Interest",
        "Penalty",
        "Fine",
        "Invoice No.",
        "Invoice Date",
        "Invoice Value",
        "Invoice Currency",
        "Exchange Rate",
        "TOI",
        "OOC",
        "Style",
        "Color",
        "Unit Price",
        "Quantity",
        "HSN Code",
        "Assessed Value",
        "BCD Rate",
        "BCD",
        "SWS Rate",
        "SWS",
        "IGST Assessed Value",
        "IGST Rate",
        "IGST",
        "GSTIN Details",
        "Freight",
        "Freight Currency",
        "Insurance",
        "Insurance Currency",
        "Misc Charges",
    ]

    newdf = newdf[column_order]

    try:
        os.remove(input_file)
    except:
        pass

    return newdf


if __name__ == "__main__":
    directory = r"C:\Users\Dhruv.jain2\OneDrive - Grant Thornton Advisory Private Limited\Documents\My Work\BOE-PDFtoExcel\New_PDFs\Incorrect Files"
    print(Extract_all(input_directory=directory))
